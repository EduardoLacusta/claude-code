"""
Importador de dados SSP-SP para o banco SQLite.

Lida com:
- Múltiplas abas por planilha (ignora metodologia, dicionário, etc.)
- Nomes de colunas diferentes entre planilhas
- Detecção automática da aba de dados principal
- Filtragem por municípios da Baixada Santista
"""
import os
import re
import logging
from datetime import datetime

import openpyxl
import requests

from app.config import SSP_SOURCES, XLSX_DIR, MUNICIPIOS_BAIXADA
from app.models import get_db, init_db

logger = logging.getLogger(__name__)

# Abas que devem ser ignoradas
SKIP_SHEETS = {"metodologia", "dicionario", "dicionário", "dicionario de dados",
               "dicionário de dados", "metadados", "leia-me", "leiame", "about", "info"}

# Mapeamento flexível de colunas — cada tipo de planilha pode ter nomes diferentes
COLUMN_MAPS = {
    "criminais": {
        "lat": ["latitude"],
        "lon": ["longitude"],
        "rubrica": ["rubrica", "especie"],
        "natureza": ["natureza_apurada", "natureza", "descr_tipolocal", "descricao"],
        "data": ["data_ocorrencia_bo", "data_ocorrencia", "dataocorrencia", "data_fato"],
        "hora": ["hora_ocorrencia_bo", "hora_ocorrencia", "horaocorrencia", "hora_fato"],
        "bairro": ["bairro", "nome_bairro"],
        "logradouro": ["logradouro", "nome_logradouro", "endereco", "logradouro_fato"],
        "municipio": ["nome_municipio_circ", "municipio_circunscricao", "municipio",
                       "cidade", "nome_municipio", "municipio_fato"],
        "tipo_local": ["descr_tipolocal", "tipo_local", "tipolocal", "local"],
    },
    "celulares": {
        "lat": ["latitude"],
        "lon": ["longitude"],
        "rubrica": ["rubrica", "natureza", "especie"],
        "data": ["data_ocorrencia_bo", "data_ocorrencia", "dataocorrencia"],
        "hora": ["hora_ocorrencia_bo", "hora_ocorrencia", "horaocorrencia"],
        "bairro": ["bairro", "nome_bairro"],
        "logradouro": ["logradouro", "nome_logradouro"],
        "municipio": ["nome_municipio_circ", "municipio_circunscricao", "municipio", "cidade"],
        "tipo_local": ["descr_tipolocal", "tipo_local", "tipolocal"],
    },
    "veiculos": {
        "lat": ["latitude"],
        "lon": ["longitude"],
        "rubrica": ["rubrica", "natureza", "especie"],
        "data": ["data_ocorrencia_bo", "data_ocorrencia", "dataocorrencia"],
        "hora": ["hora_ocorrencia_bo", "hora_ocorrencia", "horaocorrencia"],
        "bairro": ["bairro", "nome_bairro"],
        "logradouro": ["logradouro", "nome_logradouro"],
        "municipio": ["nome_municipio_circ", "municipio_circunscricao", "municipio", "cidade"],
        "tipo_local": ["descr_tipolocal", "tipo_local", "tipolocal"],
        "tipo_veiculo": ["descr_marca_veiculo", "tipo_veiculo", "marca_veiculo",
                          "descr_tipo_veiculo", "veiculo"],
    },
}


def normalize_col(name):
    """Normaliza nome de coluna para comparação."""
    if not name:
        return ""
    return re.sub(r'[^a-z0-9]', '', str(name).lower().strip())


def _sheet_looks_like_data(ws):
    """Verifica se uma aba parece conter dados reais (tem coluna 'latitude' ou muitas colunas)."""
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        cells = [str(c).strip().lower() for c in row if c is not None]
        if not cells:
            continue
        # Se tem "latitude" ou muitas colunas (>5), provavelmente é dados
        for c in cells:
            if "latitude" in c or "longitude" in c:
                return True
        if len(cells) >= 5:
            # Checa se NÃO é texto descritivo longo (típico de aba metodologia)
            if all(len(c) < 80 for c in cells):
                return True
    return False


def find_data_sheet(wb):
    """Encontra a aba principal de dados, ignorando abas de metodologia/dicionário."""
    sheets = wb.sheetnames
    if len(sheets) == 1:
        return sheets[0]

    skip_normalized = {normalize_col(s) for s in SKIP_SHEETS}

    # Primeira passada: procura aba com coluna latitude/longitude
    for name in sheets:
        if normalize_col(name) in skip_normalized:
            continue
        ws = wb[name]
        if _sheet_looks_like_data(ws):
            return name

    # Segunda passada: pega a aba com mais linhas (provavelmente é a de dados)
    best = None
    best_rows = 0
    for name in sheets:
        if normalize_col(name) in skip_normalized:
            continue
        ws = wb[name]
        rows = ws.max_row or 0
        if rows > best_rows:
            best_rows = rows
            best = name

    if best:
        return best

    return sheets[0]


def map_columns(headers, tipo):
    """Mapeia colunas do xlsx para nossos campos, por correspondência flexível."""
    col_map = COLUMN_MAPS.get(tipo, COLUMN_MAPS["criminais"])
    normalized = {normalize_col(h): i for i, h in enumerate(headers) if h}
    result = {}

    for field, candidates in col_map.items():
        for candidate in candidates:
            norm_cand = normalize_col(candidate)
            if norm_cand in normalized:
                result[field] = normalized[norm_cand]
                break

    return result


def parse_date(val):
    """Converte valor de data para YYYY-MM-DD."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s == "NULL":
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_hora(val):
    """Extrai hora como string HH:MM."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    s = str(val).strip()
    if not s or s == "NULL":
        return None
    m = re.match(r'(\d{1,2}:\d{2})', s)
    return m.group(1) if m else None


def is_baixada(municipio):
    """Verifica se o município pertence à Baixada Santista."""
    if not municipio:
        return False
    mun_upper = re.sub(r'[^A-Z ]', '', str(municipio).upper().strip())
    normalized_list = [re.sub(r'[^A-Z ]', '', m) for m in MUNICIPIOS_BAIXADA]
    return mun_upper in normalized_list


def download_xlsx(tipo, year):
    """Tenta baixar o xlsx da SSP. Retorna caminho do arquivo ou None."""
    os.makedirs(XLSX_DIR, exist_ok=True)
    src = SSP_SOURCES[tipo]
    url = src["url_tpl"].format(year=year)
    filepath = os.path.join(XLSX_DIR, src["file_tpl"].format(year=year))

    if os.path.exists(filepath):
        logger.info(f"Arquivo já existe: {filepath}")
        return filepath

    logger.info(f"Baixando {url}...")
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Referer": "https://www.ssp.sp.gov.br/",
        }
        resp = requests.get(url, headers=headers, timeout=120, stream=True)
        resp.raise_for_status()

        with open(filepath, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)

        logger.info(f"Baixado: {filepath} ({os.path.getsize(filepath)} bytes)")
        return filepath
    except Exception as e:
        logger.warning(f"Falha no download de {url}: {e}")
        # Verifica se já foi colocado manualmente
        if os.path.exists(filepath):
            return filepath
        return None


def import_xlsx(tipo, year, filepath=None):
    """
    Importa um arquivo xlsx para o banco de dados.
    Retorna dict com resultado da importação.
    """
    init_db()

    if filepath is None:
        src = SSP_SOURCES[tipo]
        filepath = os.path.join(XLSX_DIR, src["file_tpl"].format(year=year))

    if not os.path.exists(filepath):
        filepath = download_xlsx(tipo, year)
        if not filepath:
            return {"ok": False, "error": f"Arquivo não encontrado. Coloque em: data/xlsx/{SSP_SOURCES[tipo]['file_tpl'].format(year=year)}"}

    logger.info(f"Importando {filepath} (tipo={tipo}, ano={year})...")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = find_data_sheet(wb)
    ws = wb[sheet_name]
    logger.info(f"Usando aba: {sheet_name}")

    # Ler cabeçalhos — pode precisar pular linhas de "apresentação" no topo
    rows_iter = ws.iter_rows(values_only=True)
    headers = None
    col_idx = {}
    skipped_header_rows = 0

    for row in rows_iter:
        candidate = [str(h).strip() if h else "" for h in row]
        # Tenta mapear esta linha como cabeçalho
        test_idx = map_columns(candidate, tipo)
        if "lat" in test_idx and "lon" in test_idx:
            headers = candidate
            col_idx = test_idx
            break
        skipped_header_rows += 1
        if skipped_header_rows > 20:
            break

    if not headers or "lat" not in col_idx or "lon" not in col_idx:
        wb.close()
        # Mostra as primeiras linhas para debug
        return {"ok": False, "error": f"Colunas lat/lon não encontradas após {skipped_header_rows} linhas. Verifique a estrutura do arquivo."}

    logger.info(f"Colunas mapeadas (após pular {skipped_header_rows} linhas): {col_idx}")

    conn = get_db()
    # Remove dados anteriores deste tipo/ano para reimportação
    conn.execute("DELETE FROM ocorrencias WHERE tipo = ? AND ano_origem = ?", (tipo, year))

    batch = []
    count = 0
    skipped = 0

    for row in rows_iter:
        try:
            lat = row[col_idx["lat"]]
            lon = row[col_idx["lon"]]
            if lat is None or lon is None:
                skipped += 1
                continue

            lat = float(lat)
            lon = float(lon)
            if lat == 0 or lon == 0 or abs(lat) < 1 or abs(lon) < 1:
                skipped += 1
                continue

            municipio = str(row[col_idx["municipio"]]).strip().upper() if "municipio" in col_idx and row[col_idx["municipio"]] else None

            record = (
                tipo,
                lat,
                lon,
                str(row[col_idx["rubrica"]]).strip() if "rubrica" in col_idx and row[col_idx["rubrica"]] else None,
                str(row[col_idx["natureza"]]).strip() if "natureza" in col_idx and row[col_idx["natureza"]] else (str(row[col_idx["rubrica"]]).strip() if "rubrica" in col_idx and row[col_idx["rubrica"]] else None),
                parse_date(row[col_idx["data"]]) if "data" in col_idx else None,
                parse_hora(row[col_idx["hora"]]) if "hora" in col_idx else None,
                str(row[col_idx["bairro"]]).strip() if "bairro" in col_idx and row[col_idx["bairro"]] else None,
                str(row[col_idx["logradouro"]]).strip() if "logradouro" in col_idx and row[col_idx["logradouro"]] else None,
                municipio,
                str(row[col_idx["tipo_local"]]).strip() if "tipo_local" in col_idx and row[col_idx["tipo_local"]] else None,
                str(row[col_idx["tipo_veiculo"]]).strip() if "tipo_veiculo" in col_idx and row[col_idx["tipo_veiculo"]] else None,
                year,
            )
            batch.append(record)
            count += 1

            if len(batch) >= 5000:
                conn.executemany(
                    "INSERT INTO ocorrencias (tipo,lat,lon,rubrica,natureza,data,hora,bairro,logradouro,municipio,tipo_local,tipo_veiculo,ano_origem) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    batch)
                batch = []

        except (ValueError, TypeError, IndexError) as e:
            skipped += 1
            continue

    if batch:
        conn.executemany(
            "INSERT INTO ocorrencias (tipo,lat,lon,rubrica,natureza,data,hora,bairro,logradouro,municipio,tipo_local,tipo_veiculo,ano_origem) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            batch)

    # Registrar importação
    conn.execute(
        "INSERT OR REPLACE INTO import_log (tipo, ano, arquivo, registros) VALUES (?, ?, ?, ?)",
        (tipo, year, os.path.basename(filepath), count))

    conn.commit()
    conn.close()
    wb.close()

    logger.info(f"Importados {count} registros, {skipped} ignorados")
    return {"ok": True, "registros": count, "ignorados": skipped, "arquivo": os.path.basename(filepath)}


def import_all(years=None):
    """Importa todos os tipos para os anos especificados."""
    if years is None:
        years = [2025, 2026]

    results = {}
    for tipo in SSP_SOURCES:
        for year in years:
            key = f"{tipo}_{year}"
            results[key] = import_xlsx(tipo, year)
            logger.info(f"{key}: {results[key]}")

    return results


def seed_comunidades():
    """Insere/atualiza os dados fixos de comunidades/áreas de risco."""
    init_db()
    conn = get_db()

    # Sempre recriar para manter atualizado
    conn.execute("DELETE FROM comunidades")

    comunidades = [
        # === SANTOS ===
        ("Vila Gilda (Palafitas)", -23.9530, -46.3380, "SANTOS", "Maior comunidade de palafitas da América Latina"),
        ("Dique da Vila Gilda", -23.9540, -46.3370, "SANTOS", "Área de dique adjacente, sujeita a alagamentos"),
        ("Morro Nova Cintra", -23.9610, -46.3340, "SANTOS", "Uma das maiores ocupações em encosta de Santos"),
        ("Morro do Marapé", -23.9580, -46.3440, "SANTOS", "Comunidade em encosta, risco geotécnico"),
        ("Morro São Bento", -23.9650, -46.3290, "SANTOS", "Ocupação em encosta próxima ao centro histórico"),
        ("Morro Monte Serrat", -23.9670, -46.3260, "SANTOS", "Ocupações irregulares nas encostas"),
        ("Morro do Pacheco", -23.9660, -46.3240, "SANTOS", "Comunidade em encosta próxima ao centro"),
        ("Morro da Penha", -23.9640, -46.3210, "SANTOS", "Ocupação irregular, risco de deslizamento"),
        ("Morro do Fontana", -23.9590, -46.3500, "SANTOS", "Comunidade em encosta, zona noroeste"),
        ("Vila Progresso", -23.9510, -46.3420, "SANTOS", "Zona noroeste, área de mangue e alagamento"),
        ("Pantanal", -23.9490, -46.3350, "SANTOS", "Área alagadiça, ocupação sobre mangue"),
        ("Rádio Clube", -23.9470, -46.3310, "SANTOS", "Zona noroeste, sujeita a inundações"),
        ("São Manoel", -23.9520, -46.3290, "SANTOS", "Área de risco, próxima a cursos d'água"),
        ("Morro Santa Maria", -23.9630, -46.3360, "SANTOS", "Ocupação em encosta com risco geotécnico"),
        ("Caneleira", -23.9560, -46.3250, "SANTOS", "Região central/portuária, vulnerabilidade social"),

        # === SÃO VICENTE ===
        ("México 70", -23.9630, -46.3880, "S.VICENTE", "Uma das maiores comunidades, área de mangue"),
        ("Dique do Sambaiatuba", -23.9540, -46.3780, "S.VICENTE", "Palafitas às margens do Rio Sambaiatuba"),
        ("Vila Margarida (área de risco)", -23.9680, -46.3920, "S.VICENTE", "Ocupações em áreas de risco"),
        ("Saquaré / Japuí", -23.9720, -46.3950, "S.VICENTE", "Ocupação em área de mangue"),
        ("Parque das Bandeiras", -23.9570, -46.3850, "S.VICENTE", "Área de várzea sujeita a inundações"),
        ("Vila Ponte Nova", -23.9590, -46.3820, "S.VICENTE", "Área de risco próxima a córregos"),
        ("Morro dos Barbosas", -23.9610, -46.3730, "S.VICENTE", "Ocupação em encosta, risco de deslizamento"),
        ("Humaitá", -23.9480, -46.3900, "S.VICENTE", "Área continental, região de mangue"),
        ("Quarentenário", -23.9500, -46.3830, "S.VICENTE", "Área de risco próxima ao Rio Mariana"),

        # === GUARUJÁ ===
        ("Vila Baiana", -23.9870, -46.2570, "GUARUJA", "Uma das maiores comunidades, ocupação em encosta"),
        ("Morrinhos I e II", -23.9830, -46.2610, "GUARUJA", "Morro com alto risco geotécnico"),
        ("Morro do Engenho", -23.9850, -46.2540, "GUARUJA", "Área de risco em encosta"),
        ("Vila Zilda", -23.9890, -46.2550, "GUARUJA", "Ocupação densa e irregular em encosta"),
        ("Morro da Cachoeira", -23.9810, -46.2580, "GUARUJA", "Ocupação com risco de escorregamento"),
        ("Vila Áurea", -23.9860, -46.2630, "GUARUJA", "Área de vulnerabilidade social"),
        ("Morro do Macaco", -23.9820, -46.2650, "GUARUJA", "Encosta íngreme, risco classificado pela Defesa Civil"),
        ("Jardim Virgínia", -23.9910, -46.2500, "GUARUJA", "Vicente de Carvalho, ocupação irregular"),
        ("Santa Cruz dos Navegantes", -23.9950, -46.2480, "GUARUJA", "Palafitas em Vicente de Carvalho"),
        ("Jardim Progresso", -23.9930, -46.2520, "GUARUJA", "Área de risco em Vicente de Carvalho"),
        ("Morro do Outeiro", -23.9880, -46.2490, "GUARUJA", "Encosta com histórico de eventos geotécnicos"),

        # === CUBATÃO ===
        ("Cota 200", -23.8780, -46.3580, "CUBATAO", "Encosta da Serra do Mar, altíssimo risco de deslizamento"),
        ("Cota 400", -23.8720, -46.3620, "CUBATAO", "Encosta da Serra do Mar cota 400m, uma das mais perigosas do Brasil"),
        ("Cota 95/100", -23.8810, -46.3560, "CUBATAO", "Encosta da Serra do Mar, cota mais baixa"),
        ("Pilões", -23.8850, -46.3700, "CUBATAO", "Comunidade ribeirinha, enchentes e deslizamentos"),
        ("Água Fria", -23.8900, -46.3650, "CUBATAO", "Área de risco próxima a encostas e cursos d'água"),
        ("Vila Esperança", -23.8870, -46.3500, "CUBATAO", "Ocupação em área de risco"),
        ("Grotão", -23.8830, -46.3540, "CUBATAO", "Área de vale/encosta, risco de inundação"),
        ("Vila dos Pescadores", -23.9000, -46.3650, "CUBATAO", "Área de mangue às margens do estuário"),
        ("Fabril", -23.8950, -46.3600, "CUBATAO", "Ocupação irregular próxima à zona industrial"),

        # === PRAIA GRANDE ===
        ("Vila Sônia / Vila Mirim", -24.0100, -46.4300, "PRAIA GRANDE", "Ocupação irregular sujeita a alagamentos"),
        ("Jardim Quietude", -24.0050, -46.4200, "PRAIA GRANDE", "Área de risco de inundação"),
        ("Sítio do Campo", -24.0000, -46.4350, "PRAIA GRANDE", "Ocupação em várzea, sujeita a enchentes"),
        ("Jardim Samambaia", -23.9950, -46.4400, "PRAIA GRANDE", "Risco de alagamento, próxima a córregos"),
        ("Tude Bastos (área de risco)", -24.0150, -46.4150, "PRAIA GRANDE", "Bolsões de ocupação irregular"),
        ("Ribeirópolis", -24.0020, -46.4280, "PRAIA GRANDE", "Área sujeita a alagamentos"),
        ("Jardim Real", -24.0080, -46.4100, "PRAIA GRANDE", "Ocupação com riscos de alagamento"),

        # === BERTIOGA ===
        ("Indaiá (ocupações)", -23.8540, -46.1380, "BERTIOGA", "Ocupações próximas a encostas e rios"),
        ("Boracéia (ocupações)", -23.7800, -46.0200, "BERTIOGA", "Áreas de risco próximas à Serra"),
        ("Vista Linda", -23.8480, -46.1300, "BERTIOGA", "Risco geotécnico e de inundação"),
        ("Jardim Rio da Praia", -23.8520, -46.1350, "BERTIOGA", "Enchentes do Rio Itapanhaú"),
        ("Chácaras Vista Linda", -23.8500, -46.1250, "BERTIOGA", "Risco de alagamento e escorregamento"),

        # === MONGAGUÁ ===
        ("Agenor de Campos", -24.0800, -46.6200, "MONGAGUA", "Ocupações em áreas de risco"),
        ("Jardim Praia Grande (Mongaguá)", -24.0850, -46.6250, "MONGAGUA", "Área sujeita a alagamentos"),
        ("Vila Atlântica", -24.0900, -46.6180, "MONGAGUA", "Ocupação próxima a morros e córregos"),
        ("Parque Marinho", -24.0750, -46.6150, "MONGAGUA", "Ocupações irregulares sujeitas a inundação"),

        # === ITANHAÉM ===
        ("Belas Artes", -24.1800, -46.7850, "ITANHAEM", "Ocupações em região de mangue e várzea"),
        ("Jardim Suarão", -24.1750, -46.7900, "ITANHAEM", "Área sujeita a inundações"),
        ("Jardim Jamaica", -24.1850, -46.7800, "ITANHAEM", "Risco de alagamento"),
        ("Gaivota", -24.1700, -46.7950, "ITANHAEM", "Ocupações em encosta e várzea"),
        ("Cibratel", -24.1900, -46.7750, "ITANHAEM", "Áreas de risco geotécnico"),

        # === PERUÍBE ===
        ("Jardim Caraguava", -24.3100, -47.0000, "PERUIBE", "Risco de inundação e escorregamento"),
        ("Vila Peruíbe", -24.3200, -46.9950, "PERUIBE", "Ocupações em áreas de risco"),
        ("Jardim São João", -24.3150, -46.9900, "PERUIBE", "Área sujeita a alagamentos"),
        ("Balneário São João Batista", -24.3050, -46.9850, "PERUIBE", "Risco próximo a córregos e encostas"),
        ("Caraguava (Serra)", -24.3000, -47.0050, "PERUIBE", "Ocupações em encosta da Serra"),
    ]

    conn.executemany(
        "INSERT INTO comunidades (nome, lat, lon, municipio, descricao) VALUES (?,?,?,?,?)",
        comunidades)
    conn.commit()
    conn.close()
